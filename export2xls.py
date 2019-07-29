#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time     : 2019-07-27 16:28
# @Author   : Kevin Xie
# @Email    : lylwo317@gmail.com

import os
import re
from optparse import OptionParser

import openpyxl
from bs4 import BeautifulSoup
from googletrans import Translator
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from base import const
from base.utils import Utils


def add_parser():
    parser = OptionParser()

    parser.add_option("-p", "--projectDir",
                      help="Android project directory. Default value is current directory.",
                      default=".",
                      metavar="/path/to/app/directory")

    parser.add_option("-o", "--outputDir",
                      default=".",
                      help="The directory where the xlsx files will be saved. Default value is current directory.")

    parser.add_option("-t", "--googleTranslate",
                      action="store_true", default=False,
                      help="Use google translate the simplified chinese to english.")

    (options, args) = parser.parse_args()
    print("options: %s, args: %s" % (options, args))
    return options


orig_prettify = BeautifulSoup.prettify
r = re.compile(r'^(\s*)', re.MULTILINE)


def prettify(self, encoding=None, formatter="minimal", indent_width=4):
    return r.sub(r'\1' * indent_width, orig_prettify(self, encoding, formatter))


BeautifulSoup.prettify = prettify


def export_xml_to_xlsx(project_path: str):
    simplified_chinese_string_dict = {}
    for path in Utils.get_all_default_strings_xml_file_path(project_path):
        soup = BeautifulSoup(open(path, encoding='utf-8'), "xml")
        all_string_tag = soup.find_all('string')

        # 遍历所有<string>tag，将包含简体中文的字符串保存到字典
        for string_tag in all_string_tag:
            if Utils.contains_simplified_chinese(string_tag.text):
                simplified_chinese_string_dict[string_tag.attrs['name']] = string_tag.text
                print("name:", string_tag.attrs['name'], " value:", string_tag.text)

        # 遍历所有<string-array>tag，将包含简体中文的字符串保存到字典
        all_string_array_list = soup.find_all('string-array')
        for string_array_tag in all_string_array_list:
            index = 0
            for item_tag in string_array_tag.find_all('item'):
                if Utils.contains_simplified_chinese(item_tag.text):
                    key = string_array_tag.attrs['name'] + "_" + index.__str__()
                    simplified_chinese_string_dict[key] = item_tag.text
                    print("name:", key, "value:", item_tag.text)
                index += 1
    write_to_xlsx(os.path.join(options.outputDir, "strings.xlsx"), simplified_chinese_string_dict)


def write_to_xlsx(file_path: str, string_dict: dict):
    if os.path.exists(file_path):
        os.remove(file_path)
    row = 1
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Android'

    for key, value in string_dict.items():
        if row == 1:
            # title
            sheet.cell(row, 1, const.KEY_COLUMN_NAME)
            sheet.cell(row, 2, const.ZH_COLUMN_NAME)
        else:
            sheet.cell(row, 1, key)
            sheet.cell(row, 2, value)
        row += 1

    wb.save(file_path)
    print("导出未翻译的字符串到" + file_path + "。成功！")


def translate_xlsx(untranslate_xlsx_file_path):
    if not os.path.exists(untranslate_xlsx_file_path):
        return
    translator = Translator()
    wb = openpyxl.Workbook()
    sheet: Worksheet = wb.active
    sheet.title = 'Android'

    wb: Workbook = openpyxl.load_workbook(untranslate_xlsx_file_path)
    android_sheet: Worksheet = wb.get_sheet_by_name('Android')

    all_row = {}
    key_index = Utils.get_column_index_by_name_in_first_row(android_sheet, const.KEY_COLUMN_NAME) - 1
    zh_index = Utils.get_column_index_by_name_in_first_row(android_sheet, const.ZH_COLUMN_NAME) - 1
    en_column = Utils.get_column_index_by_name_in_first_row(android_sheet, const.EN_COLUMN_NAME)

    for row in android_sheet.rows:
        current_row = row[key_index].row
        android_sheet.cell(current_row, en_column, translator.translate(row[zh_index].value).text)

    wb.save(os.path.join(options.outputDir, const.EXPORT_TRANSLATE_FILE_NAME))
    return all_row


if __name__ == '__main__':
    options = add_parser()
    export_xml_to_xlsx(options.projectDir)
    if options.googleTranslate:  # 使用谷歌翻译导出的xlsx
        translate_xlsx(os.path.join(options.outputDir, const.EXPORT_XLSX_FILENAME))
